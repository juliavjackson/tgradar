import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class ExcelExporter:
    """Generates unified and styled Excel reports for Telegram Stats Bot."""

    # Unified structure for all reports (Key, Display Label, Width)
    COLUMNS = [
        ('handle', 'Канал', 20),
        ('name', 'Название канала', 30),
        ('post_link', 'Ссылка на пост', 30),
        ('post_name', 'Название поста', 30),
        ('date', 'Дата', 20),
        ('topic', 'Тематика', 15),
        ('geo', 'География', 15),
        ('subscribers', 'Подписчики', 15),
        ('views', 'Просмотры', 12),
        ('err', 'ERR %', 10),
        ('forwards', 'Репосты', 10),
        ('replies', 'Комменты', 10),
        ('reactions_pos', 'Реакции (+)', 12),
        ('reactions_neg', 'Реакции (-)', 12),
        ('price_excl_vat', 'Цена (без НДС)', 18),
        ('price_incl_vat', 'Цена (с НДС)', 18),
        ('cpv', 'CPV', 12),
        ('utm_url', 'UTM-метка', 40),
        ('visits', 'Визиты', 12),
        ('bounce_rate', 'Отказы %', 12),
        ('duration', 'Время (сек)', 15),
        ('depth', 'Глубина', 12),
    ]

    @classmethod
    def _get_styles(cls):
        return {
            'header_font': Font(name='Arial', size=11, bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid'),
            'value_font': Font(name='Arial', size=11),
            'link_font': Font(name='Arial', size=11, color='1155CC', underline='single'),
            'total_font': Font(name='Arial', size=12, bold=True),
            'total_fill': PatternFill(start_color='E2EFD9', end_color='E2EFD9', fill_type='solid'),
            'alt_fill': PatternFill(start_color='F2F7FC', end_color='F2F7FC', fill_type='solid'),
            'best_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'worst_fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'row_best_fill': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
            'row_worst_fill': PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid'),
            'thin_border': Border(
                left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
            )
        }

    @staticmethod
    def create_report(stats: dict) -> str:
        """Single post detailed report (vertical layout)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика поста"
        # ... keep original implementation for single post card ...
        # (Omitted here for brevity, but I will include it in the final write)
        return ExcelExporter._create_single_post_report(stats)

    @staticmethod
    def _create_single_post_report(stats: dict) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика поста"
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A90D9', end_color='4A90D9', fill_type='solid')
        label_font = Font(name='Arial', size=11, bold=True)
        value_font = Font(name='Arial', size=11)
        link_font = Font(name='Arial', size=11, color='1155CC', underline='single')
        section_font = Font(name='Arial', size=12, bold=True, color='4A90D9')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
        )
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 30
        ws.merge_cells('A1:B1')
        cell = ws['A1']
        cell.value = '📊 Статистика поста'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36
        row = 3
        channel = stats.get('channel', '')
        post_id = stats.get('post_id', '')
        post_url = f'https://t.me/{channel}/{post_id}'
        date_val = stats.get('date')
        date_str = date_val.strftime('%Y-%m-%d %H:%M') if isinstance(date_val, datetime) else str(date_val)
        info_items = [('Канал', f'@{channel}'), ('Пост', post_url), ('Дата публикации', date_str)]
        for label, value in info_items:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=1).border = thin_border
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.border = thin_border
            if label == 'Пост':
                val_cell.font = link_font
                val_cell.hyperlink = post_url
            else:
                val_cell.font = value_font
            row += 1
        row += 1
        ws.cell(row=row, column=1, value='📈 Метрики').font = section_font
        row += 1
        metrics = [('👁 Просмотры', stats.get('views', 0)), ('🔄 Репосты', stats.get('forwards', 0)),
                   ('💬 Комментарии', stats.get('replies', 0)), ('❤️ Реакции (всего)', stats.get('reactions_count', 0))]
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=1).border = thin_border
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.font = value_font
            val_cell.border = thin_border
            val_cell.number_format = '#,##0'
            row += 1
        filename = f"stats_{channel}_{post_id}.xlsx"
        filepath = os.path.join(tempfile.gettempdir(), filename)
        wb.save(filepath)
        return filepath

    @classmethod
    def create_campaign_report(cls, stats_list: list[dict], campaign_name: str = "Кампания") -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = campaign_name[:25]
        s = cls._get_styles()

        # Title and Date
        ws.merge_cells('A1:V1')
        ws.merge_cells('A2:V2')
        ws['A1'] = f'📊 Отчёт по кампании: {campaign_name}'
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'] = '="Дата выгрузки: "&TEXT(NOW(),"DD.MM.YYYY HH:MM")'
        ws['A2'].alignment = Alignment(horizontal='right')
        ws.row_dimensions[1].height = 32

        # Headers
        for i, (key, label, width) in enumerate(cls.COLUMNS, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            cell = ws.cell(row=3, column=i, value=label)
            cell.font = s['header_font']
            cell.fill = s['header_fill']
            cell.alignment = Alignment(horizontal='center')
            cell.border = s['thin_border']

        # Highlights detection
        cpv_vals, visit_vals, bounce_vals = [], [], []
        for i, stats in enumerate(stats_list):
            views = stats.get('views', 0)
            price = stats.get('price_excl_vat', 0)
            if views > 0 and price > 0: cpv_vals.append((i, price/views))
            m = stats.get('metrica') or {}
            if m.get('visits'): visit_vals.append((i, m['visits']))
            if m.get('bounce_rate'): bounce_vals.append((i, m['bounce_rate']))

        b_cpv = min(cpv_vals, key=lambda x: x[1])[0] if cpv_vals else -1
        w_cpv = max(cpv_vals, key=lambda x: x[1])[0] if cpv_vals else -1
        b_vis = max(visit_vals, key=lambda x: x[1])[0] if visit_vals else -1
        b_bnc = min(bounce_vals, key=lambda x: x[1])[0] if bounce_vals else -1
        w_bnc = max(bounce_vals, key=lambda x: x[1])[0] if bounce_vals else -1

        # Data rows
        DATA_START = 4
        for i, stats in enumerate(stats_list):
            r = DATA_START + i
            m = stats.get('metrica') or {}
            
            # Formulas
            err_f = f'=IF(H{r}>0,I{r}/H{r},0)' # Wait, columns changed. ERR is col 10 (J), Views is col 9 (I), Subs is col 8 (H)
            err_f = f'=IF(H{r}>0,I{r}/H{r},0)'
            cpv_f = f'=IF(I{r}>0,O{r}/I{r},"")' # CPV col 17 (Q), Views col 9 (I), Price col 15 (O)
            
            # Row mapping based on COLUMNS list
            row_data = [
                f"@{stats.get('channel', '')}",
                stats.get('channel_name', ''),
                f"https://t.me/{stats.get('channel', '')}/{stats.get('post_id', '')}" if stats.get('post_id') else "",
                stats.get('post_name', ''),
                stats.get('date').strftime('%d.%m.%Y %H:%M') if isinstance(stats.get('date'), datetime) else str(stats.get('date', '')),
                stats.get('topic', ''),
                stats.get('geo', ''),
                stats.get('subscribers', 0),
                stats.get('views', 0),
                f'=IF(H{r}>0,I{r}/H{r},0)', # ERR
                stats.get('forwards', 0),
                stats.get('replies', 0),
                stats.get('reactions_pos', 0),
                stats.get('reactions_neg', 0),
                stats.get('price_excl_vat', 0),
                stats.get('price_incl_vat', 0),
                f'=IF(I{r}>0,O{r}/I{r},"")', # CPV
                stats.get('utm_url', ''),
                m.get('visits', 0),
                m.get('bounce_rate', 0),
                m.get('duration', 0),
                m.get('depth', 0)
            ]

            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.border = s['thin_border']
                cell.font = s['link_font'] if c_idx in (3, 18) and str(val).startswith('http') else s['value_font']
                if c_idx in (3, 18) and str(val).startswith('http'): cell.hyperlink = val
                
                # Formats
                if c_idx in (8, 9, 11, 12, 13, 14, 15, 16, 19, 21): cell.number_format = '#,##0'
                elif c_idx in (10, 20): cell.number_format = '0.00%'
                elif c_idx in (17, 22): cell.number_format = '#,##0.00'

                # Colors
                if i == b_cpv: cell.fill = s['row_best_fill']
                elif i == w_cpv: cell.fill = s['row_worst_fill']
                elif r % 2 == 1: cell.fill = s['alt_fill']

                # Cell-specific
                if c_idx == 19 and i == b_vis: cell.fill = s['best_fill']; cell.font = Font(bold=True)
                if c_idx == 20:
                    if i == b_bnc: cell.fill = s['best_fill']; cell.font = Font(bold=True)
                    elif i == w_bnc: cell.fill = s['worst_fill']; cell.font = Font(bold=True)

        # Totals Row
        t_row = DATA_START + len(stats_list)
        last_r = t_row - 1
        ws.merge_cells(start_row=t_row, start_column=1, end_row=t_row, end_column=7)
        ws.cell(row=t_row, column=1, value="ИТОГО").alignment = Alignment(horizontal='right')
        
        total_formulas = {
            8: f'=SUM(H{DATA_START}:H{last_r})',
            9: f'=SUM(I{DATA_START}:I{last_r})',
            10: f'=IF(H{t_row}>0,I{t_row}/H{t_row},0)',
            11: f'=SUM(K{DATA_START}:K{last_r})',
            12: f'=SUM(L{DATA_START}:L{last_r})',
            13: f'=SUM(M{DATA_START}:M{last_r})',
            14: f'=SUM(N{DATA_START}:N{last_r})',
            15: f'=SUM(O{DATA_START}:O{last_r})',
            16: f'=SUM(P{DATA_START}:P{last_r})',
            17: f'=IF(I{t_row}>0,O{t_row}/I{t_row},"")',
            19: f'=SUM(S{DATA_START}:S{last_r})',
            20: f'=AVERAGE(T{DATA_START}:T{last_r})',
            21: f'=AVERAGE(U{DATA_START}:U{last_r})',
            22: f'=AVERAGE(V{DATA_START}:V{last_r})'
        }

        for c_idx in range(1, 23):
            cell = ws.cell(row=t_row, column=c_idx)
            cell.font = s['total_font']
            cell.fill = s['total_fill']
            cell.border = s['thin_border']
            if c_idx in total_formulas:
                cell.value = total_formulas[c_idx]
                if c_idx in (8, 9, 11, 12, 13, 14, 15, 16, 19, 21): cell.number_format = '#,##0'
                elif c_idx in (10, 20): cell.number_format = '0.00%'
                elif c_idx in (17, 22): cell.number_format = '#,##0.00'

        filename = f"{campaign_name[:20].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(path)
        return path

    @classmethod
    def create_channel_selection_report(cls, channels: list) -> str:
        """Uses the unified structure for picker results."""
        stats_list = []
        for ch in channels:
            stats_list.append({
                'channel': ch.handle,
                'channel_name': ch.name,
                'topic': ch.topic,
                'geo': ch.geo,
                'subscribers': ch.subscribers or 0,
                'views': ch.avg_reach or 0,
                'price_excl_vat': ch.price_excl_vat or 0,
                'price_incl_vat': ch.price_incl_vat or 0
            })
        return cls.create_campaign_report(stats_list, "Подборка каналов")

    @classmethod
    def create_channel_import_template(cls) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Каналы"
        s = cls._get_styles()
        for i, (key, label, width) in enumerate(cls.COLUMNS, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            cell = ws.cell(row=1, column=i, value=label)
            cell.font = s['header_font']
            cell.fill = s['header_fill']
            cell.alignment = Alignment(horizontal='center')
        
        # Example row
        examples = {
            'handle': 'digital_uz', 'name': 'Цифровой Узбекистан', 'topic': 'IT', 'geo': 'Узбекистан',
            'price_excl_vat': 500000, 'price_incl_vat': 600000
        }
        for i, (key, label, width) in enumerate(cls.COLUMNS, 1):
            if key in examples:
                cell = ws.cell(row=2, column=i, value=examples[key])
                cell.font = Font(italic=True, color='888888')
        
        path = os.path.join(tempfile.gettempdir(), "channel_import_template.xlsx")
        wb.save(path)
        return path

    @classmethod
    def create_utm_bulk_report(cls, channels: list, base_url: str, campaign_name: str) -> str:
        import re
        campaign_slug = re.sub(r'[^a-z0-9_]', '', campaign_name.lower().replace(" ", "_"))
        sep = "&" if "?" in base_url else "?"
        stats_list = []
        for ch in channels:
            utm = f"{base_url}{sep}utm_source=telegram&utm_medium={ch.handle}&utm_campaign={campaign_slug}"
            stats_list.append({
                'channel': ch.handle,
                'channel_name': ch.name,
                'topic': ch.topic,
                'geo': ch.geo,
                'price_excl_vat': ch.price_excl_vat or 0,
                'price_incl_vat': ch.price_incl_vat or 0,
                'utm_url': utm
            })
        return cls.create_campaign_report(stats_list, f"UTM_{campaign_name}")

    @classmethod
    def parse_channel_import(cls, filepath: str) -> list[dict]:
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        display_to_key = {label.lower(): key for key, label, w in cls.COLUMNS}
        raw_headers = [str(ws.cell(1, c).value).strip().lower() for c in range(1, ws.max_column + 1)]
        headers = [display_to_key.get(h, h) for h in raw_headers]
        
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            data = dict(zip(headers, row))
            handle = str(data.get('handle', '') or '').strip().lstrip('@')
            if not handle: continue
            results.append({
                'handle': handle,
                'name': str(data.get('name', handle)),
                'topic': str(data.get('topic', '') or '').strip() or None,
                'geo': str(data.get('geo', '') or '').strip() or None,
                'price_excl_vat': float(data['price_excl_vat']) if data.get('price_excl_vat') else None,
                'price_incl_vat': float(data['price_incl_vat']) if data.get('price_incl_vat') else None,
            })
        return results
